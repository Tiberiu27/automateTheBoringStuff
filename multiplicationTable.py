#!python3
#multiplicationTable.py - take an argument from command line and makes a Excel NxN.

import openpyxl, sys, os
from openpyxl.styles import Font
#move working directory to file's location
abspath = os.path.abspath(sys.argv[0])
dname = os.path.dirname(abspath)
os.chdir(dname)
if len(sys.argv) < 3:
    fontObj = Font(bold= True)
    userInput = sys.argv[1]
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(1, int(userInput)+2):
        for j in range(1, int(userInput)+ 2):
            if i == 1 and j == 1:
                sheet.cell(row= 1, column= j).value = ''
            elif i == 1:
                sheet.cell(row= i, column= j).font = fontObj
                sheet.cell(row = i, column = j).value = j -1
            elif j == 1:
                sheet.cell(row= i, column= j).font = fontObj
                sheet.cell(row= i, column= j).value = i-1
            else:
                sheet.cell(row = i, column = j).value = (i-1) * (j-1)
    

 

    wb.save('multiplicationTable.xlsx')
else:
    print('Too many arguments!')